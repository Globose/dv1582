import sqlite3
from sqlite3 import Error
from openpyxl import Workbook, load_workbook
from matplotlib import pyplot


class Analytics:
    """Class that can stora data in a database,
    and export the data into graphs and xlsx"""
    def __init__(self, path_db, table_name):
        self._connection = self._create_connection(path_db)
        self._cursor = self._connection.cursor()
        self._table_name = table_name

    def add_step(self, time: int, workers: int, products: int, food: int):
        """Add a row to the database"""
        self._cursor.execute(f"INSERT INTO {self._table_name}\
                             VALUES ({time}, {workers}, {products}, {food})")
        self._connection.commit()

    def _create_connection(self, db_file):
        """Create a sqlite3 database"""
        conn = None
        try:
            conn = sqlite3.connect(db_file)
            return conn
        except Error as e:
            print(e)
        return conn

    def create_table(self):
        """Create a table in the database"""
        sql_drop_table = "DROP TABLE IF EXISTS Resources"
        sql_create_table = f"""
            CREATE TABLE {self._table_name} (
                time INT PRIMARY KEY,
                workers INT,
                products INT,
                food INT
        );"""
        self._cursor.execute(sql_drop_table)
        self._cursor.execute(sql_create_table)

    def save_to_xlsx(self, xlsx_path: str):
        """Saves a database table to xlsx"""
        sql_select = f"SELECT * FROM {self._table_name}"
        self._cursor.execute(sql_select)
        rows = self._cursor.fetchall()

        try:
            wb = load_workbook(xlsx_path)
        except Error as e:
            print(e)
            wb = Workbook()

        if self._table_name in wb.sheetnames:
            wb.remove(wb[self._table_name])
        ws = wb.create_sheet(self._table_name)
        for row in rows:
            ws.append(row)
        wb.save(xlsx_path)

    def to_figure(self):
        """Export the database data to a graph"""
        sql_select = f"SELECT * FROM {self._table_name}"
        self._cursor.execute(sql_select)
        data = self._cursor.fetchall()

        arr_step = [row[0] for row in data]
        arr_workers = [row[1] for row in data]
        arr_products = [row[2] for row in data]
        arr_food = [row[3] for row in data]

        fig, ax = pyplot.subplots()
        ax.plot(arr_step, arr_workers, label='Workers')
        ax.plot(arr_step, arr_products, label='Products')
        ax.plot(arr_step, arr_food, label='Food')

        ax.set_title('Resources')
        ax.set_ylabel('Amount')
        ax.set_xlabel('Time (ms)')
        ax.legend()

        fig.tight_layout()
        pyplot.show()

        fig.savefig("file.png")
        pyplot.close(fig)
