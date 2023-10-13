import sqlite3
from sqlite3 import Error
from openpyxl import Workbook, load_workbook

def create_connection(db_file):
    """ create a database connection to the SQLite database
        specified by db_file
    :param db_file: database file
    :return: Connection object or None
    """
    conn = None
    try:
        conn = sqlite3.connect(db_file)
        print(sqlite3.version)
        return conn
    except Error as e:
        print(e)

    return conn


# conn = sqlite3.connect(file) # Ansluter till en databas (se nedan)
# c = conn.cursor()            # Skapar ett hjälpobjekt för databasen som används för att läsa och uppdatera tabeller.
# c.execute(sql_query)         # Skickar en fråga till databasen. Inget returneras.
# conn.commit()                # Meddelar databasen att ändringar nu kan göras. Används vid INSERT och UPDATE.
# rows = c.fetchall()          # Hämtar resultat av en förfrågan (SELECT). Vanligen som en lista med tupler.

if __name__ == '__main__':
    db_str = "C:\data\projekt\dv1582\sqlite\sqlite.db"
    conn = create_connection(db_str)
    c = conn.cursor()   #Skapa ett hjälpobjekt
    sql_drop_table = "DROP TABLE IF EXISTS Students"
    sql_create_table = """
        CREATE TABLE Students (
            Id INT PRIMARY KEY,
            Name text NOT NULL
        );"""
    sql_fill_table = "INSERT INTO Students VALUES (1,'Albert'), (2,'Robert');"
    sql_select = "SELECT * FROM Students"

    c.execute(sql_drop_table)
    c.execute(sql_create_table)
    c.execute(sql_fill_table)
    conn.commit()

    c.execute(sql_select)
    rows = c.fetchall()

    try:
        wb = load_workbook("C:\data\projekt\dv1582\sqlite\sq.xlsx")
    except:
        wb = Workbook()
    ws = wb.create_sheet("sheet")
    for row in rows:
        ws.append(row)
    wb.save("C:\data\projekt\dv1582\sqlite\sq.xlsx")
